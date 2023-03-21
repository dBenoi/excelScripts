from openpyxl import Workbook, load_workbook
import pandas as pd

netscan = 'SBPHTD032023/netScan_03152023.xlsx'
asset_list = 'SBPHTD032023/assets_032023.xlsx'

# Instantiate the workbooks
# Call the worksheet
# Remove unnecessary columns
# save and close

# netscan column removal from right to left (14,3), (4, 9), and (1, 2)
def net_col_del():
    net_wb = load_workbook(netscan)
    net_ws = net_wb.active
    net_ws.delete_cols(14, 3)
    net_ws.delete_cols(4, 9)
    net_ws.delete_cols(1, 2)
    net_wb.save(netscan)

# assets column removal from right to left (2, 8), (11, 6)
def asset_col_del():    
    asset_wb = load_workbook(asset_list)
    asset_ws = asset_wb.active
    asset_ws.delete_cols(11, 6)
    asset_ws.delete_cols(2, 8)
    asset_wb.save(asset_list)

# Sort the columns by MAC Address
def sort_netscan_by_mac():
    net_xl = pd.ExcelFile(netscan, engine='openpyxl')
    net_df = net_xl.parse(sheet_name=0)
    net_df = net_df.sort_values(by = "MAC address")
    net_df.to_excel(netscan, engine='openpyxl')

def sort_assets_by_mac():
    asset_xl = pd.ExcelFile(asset_list, engine='openpyxl')
    asset_df = asset_xl.parse(sheet_name=0)
    asset_df = asset_df.sort_values(by = "MAC Address")
    asset_df.to_excel(asset_list, engine='openpyxl')

def delete_removed():
    asset_wb = load_workbook(asset_list)
    asset_ws = asset_wb.active
    i = 1
    while i <= asset_ws.max_row:
        if asset_ws.cell(row=i, column=4).value == 'Removed':
            asset_ws.delete_rows(i, 1)
        else:
            i += 1
    asset_wb.save(asset_list)

delete_removed()






