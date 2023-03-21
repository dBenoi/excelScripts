from openpyxl.workbook import Workbook 
from openpyxl import load_workbook
import openpyxl as op

# instantiate the workbook
header = ["Mac Address", "Comments"]
new_wb = Workbook()
netscan = 'SBPHTD032023/netScan_03152023.xlsx' # network scan excel file
asset_list = 'SBPHTD032023/assets_032023.xlsx' # asset list excel file
comparison = 'SBPHTD032023/comparison_032023.xlsx' # output file of comparison

net_wb = load_workbook(netscan)
net_ws = net_wb.active

asset_wb = load_workbook(asset_list)
asset_ws = asset_wb.active

net_list = []
asset_list = []
net_unknown = []
asset_unknown = []

def net_compare():
    for net_cells in net_ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in net_cells:
                if cell.value != "00:00:00:00:00:00":
                    net_list.append(cell.value)
    
    for asset_cells in asset_ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in asset_cells:
            asset_list.append(cell.value)

    for i in net_list:
        if i not in asset_list:
            net_unknown.append(i)

    for i in asset_list:
        if i not in net_list:
            asset_unknown.append(i)

# output to new excel doc
def append_to_new():
    new_ws1 = new_wb.create_sheet("Unknown", 0)
    new_ws2 = new_wb.create_sheet("Not Found", 1)
    new_ws1.append(header)
    new_ws2.append(header)

    for r in range(0,len(net_unknown)):
        new_ws1.cell(row=r+2,column=1).value=net_unknown[r]

    for r in range(0,len(asset_unknown)):
        new_ws2.cell(row=r+2,column=1).value=asset_unknown[r]
        
    new_wb.save(filename=comparison)


def output_to_terminal():
    print("Netscan discovered the following MAC addresses that are not being tracked via the asset inventory:")
    print()
    for device in net_unknown:
        print(device)
    print()
    print("The following MAC addresses were not discovered by the network scan, but should have been:")
    print()
    for asset in asset_unknown:
        print(asset)

net_compare()
append_to_new()
output_to_terminal()